import io, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import httpx

IMAGE_SERVER_URL = os.getenv("IMAGE_SERVER_URL", "").rstrip("/")
UPLOAD_DIR = os.getenv("UPLOAD_DIR", "assets/uploads")


def fetch_image_bytes(filepath: str, preloaded: bytes | None = None) -> bytes | None:
    """Fetch image - uses preloaded bytes if provided, else local dir, else remote server."""
    if not filepath:
        return None
    if preloaded:
        return preloaded
    images_local_dir = os.getenv("IMAGES_LOCAL_DIR", "").strip()
    if images_local_dir:
        import pathlib
        local = pathlib.Path(images_local_dir) / filepath
        if local.exists():
            return local.read_bytes()
    if IMAGE_SERVER_URL:
        try:
            r = httpx.get(f"{IMAGE_SERVER_URL}/images/{filepath}", timeout=10)
            if r.status_code == 200:
                return r.content
        except Exception:
            pass
    return None


def export_to_excel(list_name: str, description: str, products: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    H_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    H_FILL  = PatternFill("solid", fgColor="1d6fb8")
    C_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin    = Side(style="thin", color="D1DCE8")
    BORD    = Border(left=thin, right=thin, top=thin, bottom=thin)
    ALT     = PatternFill("solid", fgColor="EEF4FB")
    WHITE   = PatternFill("solid", fgColor="FFFFFF")

    ws.merge_cells("A1:L1")
    ws["A1"] = f"Flyon Yiwu Market — {list_name}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="0f1f2e")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    if description:
        ws.merge_cells("A2:L2")
        ws["A2"] = description
        ws["A2"].font = Font(name="Calibri", size=10, color="7a95ae", italic=True)
        ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:L3")
    ws["A3"] = f"Total: {len(products)} products"
    ws["A3"].font = Font(name="Calibri", size=10, color="7a95ae")
    ws.row_dimensions[3].height = 16

    COLS = [
        ("Photo",        16),
        ("#",             5),
        ("Store",        18),
        ("Contact",      18),
        ("Reference",    16),
        ("Description",  28),
        ("Measurement",  14),
        ("Price (¥)",    13),
        ("QTY",           8),
        ("CBM (m³)",     10),
        ("Material",     18),
        ("Notes",        28),
    ]

    HDR = 5
    ws.row_dimensions[HDR].height = 26
    for ci, (label, width) in enumerate(COLS, 1):
        c = ws.cell(row=HDR, column=ci, value=label)
        c.font = H_FONT; c.fill = H_FILL; c.alignment = C_ALIGN; c.border = BORD
        ws.column_dimensions[get_column_letter(ci)].width = width

    for ri, p in enumerate(products, 1):
        er = HDR + ri
        fill = ALT if ri % 2 else WHITE
        ws.row_dimensions[er].height = 72

        # First photo only in column A
        image_paths = p.get("image_paths", [])
        if isinstance(image_paths, str):
            image_paths = [x.strip() for x in image_paths.split(",") if x.strip()]

        first_path = image_paths[0] if image_paths else None
        if first_path:
            img_bytes = fetch_image_bytes(first_path)
            if img_bytes:
                try:
                    import io as _io
                    img = XLImage(_io.BytesIO(img_bytes))
                    img.width = 75; img.height = 68
                    ws.add_image(img, f"A{er}")
                except Exception:
                    ws.cell(row=er, column=1, value="[img]")
        ws.cell(row=er, column=1).fill = fill
        ws.cell(row=er, column=1).border = BORD

        values = [
            ri,
            p.get("store", ""),
            p.get("store_contact", ""),
            p.get("reference", ""),
            p.get("description", ""),
            p.get("measurement", ""),
            p.get("price"),
            p.get("qty"),
            p.get("cbm"),
            p.get("material", ""),
            p.get("notes", ""),
        ]
        for ci, val in enumerate(values, 2):
            cell = ws.cell(row=er, column=ci, value=val if val is not None else "")
            cell.fill = fill; cell.border = BORD
            cell.alignment = C_ALIGN if ci in (2, 8, 9, 10) else L_ALIGN
            if ci == 8 and isinstance(val, (int, float)):
                cell.number_format = '"¥"#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
